"""
Excel导出模块
============
将 results/*.csv 汇总导出为 results/全部结果.xlsx，每道题一个 sheet。
"""

import os
import sys
import warnings

# 确保可以从 utils/ 目录内导入同级的 constants 和根目录的 utils 包
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.constants import (
    ALK_H2_RATE, ALK_RATED_POWER,
    BASE_LOAD_PEAK,
    H2_PER_TON_NH3, HOURS_PER_DAY,
    NH3_RATED_POWER,
    PEM_H2_RATE, PEM_RATED_POWER,
    PV_CAPACITY,
    RATED_DAILY_NH3,
    STORAGE_INV_COST, STORAGE_LIFE_YEARS,
    WIND_CAPACITY,
)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _get_results_dir():
    """获取 results 目录的绝对路径（相对于本脚本位置）。"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, "..", "results")


def _read_csv_or_none(filename):
    """读取 CSV 文件，不存在时返回 None。"""
    results_dir = _get_results_dir()
    path = os.path.join(results_dir, filename)
    if not os.path.exists(path):
        warnings.warn(f"CSV 文件不存在，跳过: {path}")
        return None
    return pd.read_csv(path, encoding="utf-8-sig")



def _auto_fit_columns(ws):
    """根据内容自动调整列宽。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                # 中文字符按 2 倍宽度计算
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
            except (TypeError, ValueError, AttributeError):
                # column width adjustment failure is non-critical, use default width
                pass
        # 限制最大宽度 60，最小宽度 8
        adjusted = min(max_len + 3, 60)
        adjusted = max(adjusted, 8)
        ws.column_dimensions[col_letter].width = adjusted


def _format_sheet(ws):
    """为工作表添加基本格式：表头加粗、冻结首行、自动列宽。"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 数据区域加边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _write_df_to_sheet(ws, df, start_row=1):
    """将 DataFrame 写入工作表（含表头）。"""
    # 写表头
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)
    # 写数据
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def export_all():
    """主导出函数：读取所有 CSV 并写入一个 XLSX 文件。"""
    results_dir = _get_results_dir()
    output_path = os.path.join(results_dir, "全部结果.xlsx")
    os.makedirs(results_dir, exist_ok=True)

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    sheet_defs = [
        ("Q1场景结果", "q1_results.csv",     True),   # 读取场景结果 CSV
        ("Q2结果",   "q2_results.csv",        True),   # 直接读取
        ("Q3结果",   "q3_results.csv",         True),   # 需与 cluster 合并
        ("Q4结果",   "q4_results.csv",         True),   # 直接读取
        ("Q5敏感性", "q5_sensitivity.csv",     True),   # 直接读取
    ]

    row_counts = {}

    for sheet_name, csv_name, is_direct in sheet_defs:
        print(f"正在导出 {sheet_name}...")
        df = _read_csv_or_none(csv_name)
        if df is None:
            continue

        if sheet_name == "Q3结果":
            cluster_df = _read_csv_or_none("q3_cluster_labels.csv")
            if cluster_df is not None and "场景编号" in df.columns and "场景编号" in cluster_df.columns:
                # Coerce both to int to avoid str/int merge conflict (summary row makes q3_results.csv a str column)
                df["场景编号"] = pd.to_numeric(df["场景编号"], errors="coerce")
                cluster_df["场景编号"] = pd.to_numeric(cluster_df["场景编号"], errors="coerce")
                df = df.merge(cluster_df, on="场景编号", how="left")

        ws = wb.create_sheet(title=sheet_name)
        _write_df_to_sheet(ws, df)
        _format_sheet(ws)
        row_counts[sheet_name] = len(df)

    wb.save(output_path)

    # 文件大小
    file_size_kb = os.path.getsize(output_path) / 1024
    print(f"\nExcel 导出完成: results/全部结果.xlsx ({file_size_kb:.1f} KB, {len(wb.sheetnames)} 个工作表)")
    for name in wb.sheetnames:
        print(f"  {name}: {row_counts.get(name, '?')} 行")


if __name__ == "__main__":
    export_all()
